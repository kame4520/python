import pathlib  #標準ライブラリ
import openpyxl #外部ライブラリ
import csv      #標準ライブラリ
import os       #標準ライブラリ
import sys

"""
 (差分)シートに特別パッチが必要となる日が存在しないかチェックして結果をresultに入れる
"""
def spc_patch_chk(sheet_name):
	result=[]
	ws = wb[sheet_name]
	max_R=ws.max_row
	max_C=ws.max_column
	cnt = 0
	print(sheet_name+' の処理を開始')
	for i in range(4,max_C):
		chk_cell = i -1
		for j in range(6,max_R):
			if ws.cell(row=j, column=chk_cell).value - ws.cell(row=j, column=i).value >= 2:
				result.append(
				str(format(ws.cell(row=1,column=i).value,'%Y/%m/%d %H:%M:%S')) +'\t'
				+ ws.cell(row=j,column=1).value +'\t'
				+ prfc[str(ws.cell(row=j, column=1).value)] +'\t'
				+ str(ws.cell(row=j, column=chk_cell).value - ws.cell(row=j, column=i).value)
				+ '\n'
				)
				cnt += 1
	print(sheet_name+' の処理を完了\n\n')
	if cnt == 0:
		print(sheet_name+' には特別パッチが必要なレコードはありませんでした。\n\n')
	return result

"""
都道府県コードを辞書に登録
"""
prfc={"北海道":"01","青森":"02","岩手":"03","宮城":"04","秋田":"05","山形":"06","福島":"07","茨城":"08","栃木":"09","群馬":"10","埼玉":"11",
	"千葉":"12","東京":"13","神奈川":"14","新潟":"15","富山":"16","石川":"17","福井":"18","山梨":"19","長野":"20","岐阜":"21","静岡":"22",
	"愛知":"23","三重":"24","滋賀":"25","京都":"26","大阪":"27","兵庫":"28","奈良":"29","和歌山":"30","鳥取":"31","島根":"32","岡山":"33",
	"広島":"34","山口":"35","徳島":"36","香川":"37","愛媛":"38","高知":"39","福岡":"40","佐賀":"41","長崎":"42","熊本":"43","大分":"44",
	"宮崎":"45","鹿児島":"46","沖縄":"47"}


"""処理はここから------------------------------------------------------"""
"""
作業ディレクトリの確認
今はデスクトップ直下の「python￥wrk」で固定にしてる。
"""
print("特別パッチの要否チェックを開始します。\n")
desktop_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Desktop"
kakunin_dir = desktop_path+'\python\wrk'
if os.path.exists(kakunin_dir) == False:
    print("作業ディレクトリ "+kakunin_dir+" が存在しません。\n処理を終了します。")
    sys.exit(1)
os.chdir(kakunin_dir)
wrk_dir = os.getcwd()

#print(wrk_dir)

"""
SQLの材料フィルを見つけて開く
利用しているopenpexlではxlsxファイルじゃないと扱えないので拡張子は固定
"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    if pass_obj.match("*D+N日表記*.xlsx"):
            wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする
print(str(pass_obj) + " に対して処理を行います。\n\n")


"SEJ店舗のチェック"
with open('SEJ_SPCL_PATCH_CHK.csv', 'w', encoding = 'utf_8', newline = '\n') as sej_sp:
	writer = csv.writer(sej_sp)
	for initem in spc_patch_chk('ＳＥＪ店舗 (差分)'):
		sej_sp.write(initem)
	sej_sp.close()

"その他のチェック"
with open('OTH_SPCL_PATCH_CHK.csv', 'w', encoding = 'utf_8', newline = '\n') as oth_sp:
	writer = csv.writer(oth_sp)
	for initem in spc_patch_chk('その他 (差分)'):
		oth_sp.write(initem)
	oth_sp.close()

print("特別パッチの要否チェックは全て完了しました。")
#print("処理件数は"+str(cnt)+"件です\n\n")
