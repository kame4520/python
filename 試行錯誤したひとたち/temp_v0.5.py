import pathlib  #標準ライブラリ
import openpyxl #外部ライブラリ
import csv      #標準ライブラリ
import os       #標準ライブラリ
import pprint   #外部ライブラリ 結果を見やすくできるっぽいので入れてみた

"""
作業ディレクトリの確認
今はデスクトップ直下の「python￥wrk」で固定にしてる。
"""
os.chdir('.\\python\wrk')
wrk_dir = os.getcwd()
print(wrk_dir)


"""
SQLの材料フィルを見つける
利用しているopenpexlではxlsxファイルじゃないと扱えないので拡張子は固定
"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    if pass_obj.match("*D+N日表記*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする

"""
対象のシートから休配リードタイムを読み込む関数
"""
def READ_LDTM():
    result =[]
    global cnt
    for i in range(3,max_C+1):
        for k in range(6,max_R+1):
            if  str(ws.cell(row=k, column=i).value) != '0' :
                result.append(
                'TO_DATE(\''+
                str(format(ws.cell(row=1, column=i).value,'%Y/%m/%d %H:%M:%S')) +
                ',\'yyyy/mm/dd hh24:mi:ss\')'+
                ','+ str(ws.cell(row=k, column=1).value) +
                ','+ str(ws.cell(row=k, column=i).value) +
                '\n'
                )
                cnt += 1
    return result

"""
「ＳＥＪ店舗 (差分)」シートから休配リードタイムを取得
 リードタイムが"0"でないセルだけを処理の対象とする。
"""
ws = wb.worksheets[8] # シート名で指定したいので後で調べる
max_R=ws.max_row
max_C=ws.max_column
cnt = 0
print(ws.title+"シートの処理を開始")
with open('SEJ_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as sej:
    writer = csv.writer(sej)
    for initem in READ_LDTM():
        sej.write(initem)
    sej.close()
print(ws.title+"シートの処理が終了しました。")
print("処理件数は"+str(cnt)+"件です\n\n")

"""
「その他 (差分)」シートから休配リードタイムを取得
 リードタイムが"0"でないセルだけを処理の対象とする。
"""
ws = wb.worksheets[9] # シート名で指定したいので後で調べる
max_R=ws.max_row
max_C=ws.max_column
cnt = 0
print(ws.title+"シートの処理を開始")
with open('OTHER_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as oth:
    writer = csv.writer(oth)
    for initem in READ_LDTM():
        oth.write(initem)
    oth.close()
print(ws.title+"シートの処理が終了しました。")
print("処理件数は"+str(cnt)+"件です")
