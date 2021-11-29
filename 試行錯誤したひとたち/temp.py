import pathlib  #標準ライブラリ
import openpyxl #外部ライブラリ
import csv      #標準ライブラリ
import os       #標準ライブラリ
import pprint   #外部ライブラリ 結果を見やすくできるっぽいので入れてみた
import numpy as np

os.chdir('.\\python\wrk')
wrk_dir = os.getcwd()
print(wrk_dir)

"""
SQLの材料フィルを見つける
利用しているopenpexlではxlsxファイルじゃないと扱えないので拡張子は固定
"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    if pass_obj.match("*D+N日表記*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする
print(pass_obj)
print(wb.sheetnames)

"""
対象のシートから休配リードタイムを読み込む関数
"""
def READ_LDTM():
    result =[]
    for i in range(3,max_C+1):
        for k in range(6,max_R+1):
            if  str(ws.cell(row=k, column=i).value) != '0' :

                result.append(
                str(format(ws.cell(row=1, column=i).value,'%Y/%m/%d')) +
                ','+ str(ws.cell(row=k, column=1).value) +
                ','+ str(ws.cell(row=k, column=i).value)
                )

    return result

"""
「ＳＥＪ店舗 (差分)」シートから休配リードタイムを取得
 リードタイムが"0"でないセルだけを処理の対象とする。
"""
ws = wb.worksheets[8] # シート名で指定したいので後で調べる
max_R=ws.max_row
max_C=ws.max_column
print(wb.active.title+"シートの処理を開始")
with open('SEJ_INSERT.csv', 'w', encoding = 'utf_8', newline = '\r\n') as sej:
    writer = csv.writer(sej)

#    for w in len(READ_LDTM()) -1
#        sej.write(str(READ_LDTM()))
    print(READ_LDTM())
    R = READ_LDTM()
    for item in R:
        print(item)
        sej.write(item)
    #sej.write(str(READ_LDTM()))
    sej.close()
print(wb.active.title+"シートの処理が終了しました")

"""
「その他 (差分)」シートから休配リードタイムを取得
 リードタイムが"0"でないセルだけを処理の対象とする。
"""
ws = wb.worksheets[9] # シート名で指定したいので後で調べる
max_R=ws.max_row
max_C=ws.max_column
print(wb.active.title+"シートの処理を開始")
with open('OTHER_INSERT.csv', 'w', encoding = 'utf_8', newline = '\r\n') as oth:
    writer = csv.writer(oth)
    oth.write(str(READ_LDTM()))
    oth.close()
print(wb.active.title+"シートの処理が終了しました")
#with open(path\'SEJ_INSERT.csv', 'w') as SEf:
