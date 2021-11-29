import pathlib  #標準ライブラリ
import os       #標準ライブラリ
import openpyxl #外部ライブラリ

"""A列の「都道府県」に入っているスペースを除去する関数"""
def RMV_SPACE():
    max_R=ws.max_row
    for i in range(6,max_R):
            ws.cell(row=i,column=1).value = str(ws.cell(row=i,column=1).value).strip()

"""新潟修正用の関数"""
def NIGATA_MODIFY():
    max_R=ws.max_row
    max_C=ws.max_column
    for j in range(6,max_R):
            if ws.cell(row=j,column=1).value == '新潟':
                nigata = j
            elif ws.cell(row=j,column=1).value == '富山':
                toyama = j
    for l in range(3,max_C):
        tym = ws.cell(row=toyama,column=l).value
        ws.cell(row=nigata,column=l,value = tym)

"""差分シートとD+Nシート用の関数"""
def KANSU():
    max_R=ws.max_row
    max_C=ws.max_column
    for m in range(6,max_R):
            if ws.cell(row=m,column=1).value == '北海道':
                hokkaido = m
    for n in range(6,max_R+1):
        for o in range(3,max_C+1):
            if kansu == 'sej_sabun':
                tgt1 = str(ws.cell(row=n,column=o).coordinate)
                tgt2 = str("""'ＳＥＪ店舗(基表) (新潟修正)'""")
                tgt3 = str("""'ＳＥＪ店舗 (新潟修正)'""")
                ws.cell(row=n,column=o,value = sabun_kansu.format(tgt1,tgt2,tgt3))
            elif kansu == 'oth_sabun':
                tgt1 = str(ws.cell(row=n,column=o).coordinate)
                tgt2 = str("""'その他（基表） (新潟修正)'""")
                tgt3 = str("""'その他 (新潟修正)'""")
                ws.cell(row=n,column=o,value = sabun_kansu.format(tgt1,tgt2,tgt3))
            elif kansu == 'sej_dn':
                tgt1 = str(ws.cell(row=n,column=o).coordinate)
                tgt2 = str("""'ＳＥＪ店舗 (差分)'""")
                tgt3 = str("""'ＳＥＪ店舗(基表) (新潟修正)'""")
                ws.cell(row=n,column=o,value = dn_kansu.format(tgt1,tgt2,tgt3))
            elif kansu == 'oth_dn':
                tgt1 = str(ws.cell(row=n,column=o).coordinate)
                tgt2 = str("""'その他 (差分)'""")
                tgt3 = str("""'その他（基表） (新潟修正)'""")
                ws.cell(row=n,column=o,value = dn_kansu.format(tgt1,tgt2,tgt3))

#「差分」用のエクセル関数
sabun_kansu = """=IF(DATE(MID({1}!{0},1,4),MID({1}!{0},5,2),MID({1}!{0},7,2))<DATE(MID({2}!{0},1,4),\
MID({2}!{0},5,2),MID({2}!{0},7,2)),DATEDIF(DATE(MID({1}!{0},1,4),MID({1}!{0},5,2),\
MID({1}!{0},7,2)),DATE(MID({2}!{0},1,4),MID({2}!{0},5,2),MID({2}!{0},7,2)),  "D"),)"""
#「D+N」用のエクセル関数
dn_kansu = """=IF({1}!{0}=0,{2}!{0},{2}!{0}&"+"&{1}!{0})"""

"""
作業ディレクトリの確認
今はデスクトップ直下の「python\wrk\」で固定にしてる。
ある程度できたらinput使ってプロンプト上で入力するかファイルを読み込むか...
"""
desktop_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Desktop"
os.chdir(desktop_path+'\python\wrk')
wrk_dir = os.getcwd()
print(wrk_dir +" 配下のファイルに対して処理を実行します。\n")

"""材料となるファイル名とパスを取得 多分ちゃんと条件になってない。。。。"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    pass_obj.match("店舗留置き*店着スケジュール*.xlsx")
    print("ファイル名を確認。")

wb = openpyxl.load_workbook(pass_obj)

""" ファイルを別名保存 """
dname = os.path.dirname(pass_obj)
org_fname = os.path.basename(pass_obj)
org_fname1 = os.path.splitext(os.path.basename(pass_obj))[0]
fname = dname+ "\\" + org_fname1 +"（D+N日表記）"+".xlsx"
print(org_fname+"　を加工します。\n\n")
wb.save(fname)

""" シート名の変更 & 都道府県に入っているスペースを除去 """
print("シート名の更新を開始")
ws = wb['ＳＥＪ店舗']
RMV_SPACE()

ws = wb['ＳＥＪ店舗(基表)']
RMV_SPACE()

ws = wb['AH、SS、IY、LOFT、YB、YMT、SG、デニーズ店舗']
ws.title = 'その他'
RMV_SPACE()

ws = wb['AH、SS、IY、LOFT、YB、YMT、SG、デニーズ（基表']
ws.title = 'その他（基表）'
RMV_SPACE()
print("シート名の更新が完了\n")


""" 新潟修正シートの用意"""
print("「新潟修正」シートの作成開始")

ws = wb['ＳＥＪ店舗']
wb.copy_worksheet(ws)
ws = wb.worksheets[4]
ws.title = 'ＳＥＪ店舗 (新潟修正)'
NIGATA_MODIFY()
ws.sheet_view.zoomScale = 85

ws = wb['その他']
wb.copy_worksheet(ws)
ws = wb.worksheets[5]
ws.title = 'その他 (新潟修正)'
NIGATA_MODIFY()
ws.sheet_view.zoomScale = 85

ws = wb['ＳＥＪ店舗(基表)']
wb.copy_worksheet(ws)
ws = wb.worksheets[6]
ws.title = 'ＳＥＪ店舗(基表) (新潟修正)'
NIGATA_MODIFY()
ws.sheet_view.zoomScale = 85

ws = wb['その他（基表）']
wb.copy_worksheet(ws)
ws = wb.worksheets[7]
ws.title = 'その他（基表） (新潟修正)'
NIGATA_MODIFY()
ws.sheet_view.zoomScale = 85
print("「新潟修正」シートの作成完了\n")



""" 差分シートの作成"""
print("「差分」シートの作成開始")
ws = wb['ＳＥＪ店舗 (新潟修正)']
wb.copy_worksheet(ws)
ws = wb.worksheets[8]
ws.title = 'ＳＥＪ店舗 (差分)'
ws.sheet_view.zoomScale = 85
kansu = 'sej_sabun'
KANSU()

ws = wb['その他 (新潟修正)']
wb.copy_worksheet(ws)
ws = wb.worksheets[9]
ws.title = 'その他 (差分)'
ws.sheet_view.zoomScale = 85
kansu = 'oth_sabun'
KANSU()
print("「差分」シートの作成完了\n")

""" D+Nシートの作成"""
print("「D+N」シートの作成開始")
ws = wb['ＳＥＪ店舗 (新潟修正)']
wb.copy_worksheet(ws)
ws = wb.worksheets[10]
ws.title = 'ＳＥＪ店舗 (D+N)'
ws.sheet_view.zoomScale = 85
kansu = 'sej_dn'
KANSU()


ws = wb['その他 (新潟修正)']
wb.copy_worksheet(ws)
ws = wb.worksheets[11]
ws.title = 'その他 (D+N)'
ws.sheet_view.zoomScale = 85
kansu = 'oth_dn'
KANSU()
print("「D+N」シートの作成完了")

wb.save(fname)
