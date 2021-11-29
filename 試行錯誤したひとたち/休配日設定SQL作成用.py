import pathlib  #標準ライブラリ
import openpyxl #外部ライブラリ
import csv      #標準ライブラリ
import os       #標準ライブラリ
import pprint   #外部ライブラリ 結果を見やすくできるっぽいので入れてみた

os.chdir('.\\python\wrk')
wrk_dir = os.getcwd()
print(wrk_dir)

"""
SQLの材料フィルを見つける
利用しているopenpexlではxlsxファイルじゃないと扱えないので拡張子は固定
"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    if pass_obj.match("*D+N日表記*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする
print(pass_obj)
print(wb.sheetnames)

"""
「ＳＥＪ店舗 (差分)」シートから休配リードタイムを取得
 リードタイムが"0"でないセルだけを処理の対象とする。
"""
ws = wb.worksheets[8] # シート名で指定したいので後で調べる
max_R=ws.max_row
max_C=ws.max_column
print(ws)
print(max_R)
print(max_C)
print(wb.active.title+"シートの処理を開始")
for i in range(3,max_C+1):
    for k in range(6,max_R+1):
        if  str(ws.cell(row=k, column=i).value) != '0' :
            print(
            'TO_DATE(\''+
            str(format(ws.cell(row=1, column=i).value,'%Y/%m/%d') +
            ',\'yyyy/mm/dd hh24:mi:ss\')'+
            ','+ str(ws.cell(row=k, column=1).value) +
            ','+ str(ws.cell(row=k, column=i).value)
            ))


#with open(path\'SEJ_INSERT.csv', 'w') as SEf:
