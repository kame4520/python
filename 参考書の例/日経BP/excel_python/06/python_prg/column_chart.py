import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("..\data\column_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=3, max_col=3, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #横棒
# 横棒にすると、得意先名の一部が消えるので、chart全体の高さを指定する
#chart.height = 10

chart.style = 28    #1グレイ、11青、28だいだい、30黄色、37グラフの背景が薄いグレイ、45全体の背景が黒
chart.title = "得意先別売上"
chart.y_axis.title = "売上額"
chart.x_axis.title = "得意先名"

chart.add_data(data,titles_from_data=True)  #当月売上が凡例になる
chart.set_categories(labels)
sh.add_chart(chart, "E3")

wb.save("..\data\column_chart.xlsx")