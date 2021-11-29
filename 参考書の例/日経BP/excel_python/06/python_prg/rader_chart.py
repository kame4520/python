import openpyxl
from openpyxl.chart import RadarChart, Reference

wb = openpyxl.load_workbook(r"..\data\radar_chart.xlsx")
sh = wb.active

data = Reference(sh, min_col=2, max_col=4, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = RadarChart()
#デフォルトはstandard
#filledは塗りつぶし
#chart.type = "filled"
chart.title = "部門別売上"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "F2")
wb.save(r"..\data\radar_chart.xlsx")