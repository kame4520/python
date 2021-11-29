import openpyxl
from openpyxl.chart import LineChart, Reference

wb = openpyxl.load_workbook("..\data\line_chart.xlsx")
sh = wb.active

data = Reference(sh, min_col=2, max_col=9, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = LineChart()
chart.title = "月別売上"
chart.y_axis.title = "売上数量"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "A9")
wb.save("..\data\line_chart.xlsx")