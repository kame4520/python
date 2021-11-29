import openpyxl

wb = openpyxl.Workbook()
sh = wb.active

sh["b2"] = "セルの結合をテスト"
sh.merge_cells("b2:c2")
sh["b2"].alignment = openpyxl.styles.Alignment(horizontal="center")
#sh.unmerge_cells("b2:c2")

wb.save(r"..\data\format_test.xlsx")