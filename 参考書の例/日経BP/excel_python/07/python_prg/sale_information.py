from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm
import openpyxl 
import pathlib  
import datetime
from PIL import Image

def load_informatiom():
    wb = openpyxl.load_workbook("..\data\セールの案内.xlsx")
    sh = wb.active
    sale_dict = {} 
    for row in range(1, sh.max_row + 1):
        if sh.cell(row,1).value == "案内文":
            info_list = [sh.cell(row,2).value]
            for info_row in range(row + 1 , sh.max_row + 1):
                info_list.append(sh.cell(info_row,2).value)
            sale_dict.setdefault("案内文", info_list)
        elif sh.cell(row,1).value is not None:     
            sale_dict.setdefault(sh.cell(row,1).value, sh.cell(row,2).value)
    return sale_dict


sale_dict = load_informatiom()
path = pathlib.Path("..\data\sales\pdf")
wb = openpyxl.load_workbook("..\data\得意先台帳.xlsx")
sh = wb["宛名書き"]
for row in range(1, sh.max_row + 1):
    file_name = (sh.cell(row,2).value) + "様ご案内.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("セール案内")
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawCentredString(6*cm, 27*cm, sh.cell(row,2).value + " " \
        + sh.cell(row,3).value + " 様")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm) #得意先名の下にライン
    cv.setFont("HeiseiKakuGo-W5", 14)
    cv.drawCentredString(10*cm, 24*cm, sale_dict["タイトル"])
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawString(2*cm, 22*cm, "開催日時：" + sale_dict["開催日時"])
    cv.drawString(2*cm, 21*cm, "開催場所：" + sale_dict["開催場所"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HeiseiKakuGo-W5", 12)
    for line in sale_dict["案内文"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14) # POSITIVE Y moves down!!!
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    image =Image.open("..\data\logo.png")
    cv.drawInlineImage(image,13*cm,13*cm)
    cv.showPage()
    cv.save()