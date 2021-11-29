import os
import sys
import csv
import pathlib

import openpyxl #外部ライブラリ
import PySimpleGUI as sg #外部ライブラリ


"""
対象のシートから休配リードタイムを読み込む関数
"""
def READ_LDTM():
    result =[]
    global cnt
    if ws.title == "ＳＥＪ店舗 (差分)":
        uktr_flg = "0"
    else:
        uktr_flg = "1"
    for i in range(3,max_C+1):
        for k in range(6,max_R+1):
            if  str(ws.cell(row=k, column=i).value) != '0' :
                result.append(
                'TO_DATE(\'' + str(format(ws.cell(row=1, column=i).value,'%Y/%m/%d %H:%M:%S')) + ',\'yyyy/mm/dd hh24:mi:ss\')'#作業日
                +'\t'+ str(ws.cell(row=k, column=1).value)  #都道府県名
                +'\t'+ prfc[str(ws.cell(row=k, column=1).value)]  #都道府コード
                +'\t'+ uktr_flg                             #受取事業会社FLG
                +'\t'+ str(ws.cell(row=k, column=i).value)  #休配リードタイム
                +'\t'+ '\n'
                )
                cnt += 1
    return result

"""
都道府県コードを辞書に登録
"""
prfc={"北海道":"01","青森":"02","岩手":"03","宮城":"04","秋田":"05","山形":"06","福島":"07","茨城":"08","栃木":"09","群馬":"10","埼玉":"11",
	"千葉":"12","東京":"13","神奈川":"14","新潟":"15","富山":"16","石川":"17","福井":"18","山梨":"19","長野":"20","岐阜":"21","静岡":"22",
	"愛知":"23","三重":"24","滋賀":"25","京都":"26","大阪":"27","兵庫":"28","奈良":"29","和歌山":"30","鳥取":"31","島根":"32","岡山":"33",
	"広島":"34","山口":"35","徳島":"36","香川":"37","愛媛":"38","高知":"39","福岡":"40","佐賀":"41","長崎":"42","熊本":"43","大分":"44",
	"宮崎":"45","鹿児島":"46","沖縄":"47"}



#  セクション1 - オプションの設定と標準レイアウト
sg.theme('Dark Blue 3')

layout = [
    [sg.Text('全ての項目を入力してください')],
    [sg.Text('ファイル保管場所', size=(15, 1)), sg.InputText('C:\\Users\\kamedai\\Desktop\\python\\wrk')],
    [sg.Text('ファイル名', size=(15, 1)), sg.InputText('店舗留置き　店着スケジュール202003（D+N日表記）.xlsx')],
    [sg.Submit(button_text='実行ボタン')]
]

# セクション 2 - ウィンドウの生成
window = sg.Window('トーハンお届け予定日　作業ファイル作成', layout)

# セクション 3 - イベントループ
while True:
    event, values = window.read()

    if event is None:
        print('exit')
        break

    if event == '実行ボタン':
        wrk_dir   = values[0]
        work_file = values[1]
        pass_obj  = wrk_dir+'\\'+work_file
        if os.path.exists(wrk_dir) == False:
            show_message = print("指定されたディレクトリが存在しません。\n処理を終了します。")
            sys.exit(1)
        elif os.path.exists(wrk_dir+'\\'+work_file) == False:
            show_message = print("指定されたファイルが存在しません。\n処理を終了します。")
            sys.exit(1)
        else:
            show_message = os.path.exists(wrk_dir)

        wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする
        #show_message += str(pass_obj + " に対して処理を行います。\n\n")

        """
        「ＳＥＪ店舗 (差分)」シートから休配リードタイムを取得
         リードタイムが"0"でないセルだけを処理の対象とする。
        """
        ws = wb['ＳＥＪ店舗 (差分)']
        max_R=ws.max_row
        max_C=ws.max_column
        cnt = 0
        #show_message += print(ws.title+"シートの処理を開始")
        with open('SEJ_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as sej:
            writer = csv.writer(sej)
            for initem in READ_LDTM():
                sej.write(initem)
            sej.close()
        #show_message += print(ws.title+"シートの処理が終了しました。")
        #show_message += print("処理件数は"+str(cnt)+"件です\n\n")

        """
        「その他 (差分)」シートから休配リードタイムを取得
        リードタイムが"0"でないセルだけを処理の対象とする。
        """
        ws = wb['その他 (差分)'] # シート名で指定したいので後で調べる
        max_R=ws.max_row
        max_C=ws.max_column
        cnt = 0
        #print(ws.title+"シートの処理を開始")
        with open('OTHER_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as oth:
            writer = csv.writer(oth)
            for initem in READ_LDTM():
                oth.write(initem)
            oth.close()
    #    show_message += print(ws.title+"シートの処理が終了しました。")
    #    show_message += print("処理件数は"+str(cnt)+"件です")

    print(show_message)
        # ポップアップ
    sg.popup(show_message)


# セクション 4 - ウィンドウの破棄と終了
window.close()
