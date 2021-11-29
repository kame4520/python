import pathlib  #標準ライブラリ
import openpyxl #外部ライブラリ
import csv      #標準ライブラリ
import os       #標準ライブラリ
import sys

"""
対象のシートから休配リードタイムを読み込む関数
"""
def READ_LDTM(sheet_name):
    result =[]
    ws = wb[sheet_name]
    print(sheet_name+' の処理を開始')
    global cnt
    max_R=ws.max_row
    max_C=ws.max_column
    cnt = 0
    if ws.title == "ＳＥＪ店舗 (差分)":
        uktr_flg = "0"
    else:
        uktr_flg = "1"
    for i in range(3,max_C):
        for k in range(6,max_R):
            if  str(ws.cell(row=k, column=i).value) != '0' :
                result.append(
                'TO_DATE(\'' + str(format(ws.cell(row=1, column=i).value,'%Y/%m/%d %H:%M:%S')) + ',\'yyyy/mm/dd hh24:mi:ss\')'#作業日
                +'\t'+ str(ws.cell(row=k, column=1).value)  #都道府県名
                +'\t'+ prfc[str(ws.cell(row=k, column=1).value)]  #都道府コード
                +'\t'+ uktr_flg                             #受取事業会社FLG
                +'\t'+ str(ws.cell(row=k, column=i).value)  #休配リードタイム
                +'\t'+ '\n'
                )
                cnt += 1
    print(sheet_name +' の処理を完了')
    print(sheet_name+' の処理結果は'+ str(cnt) +'件でした。\n')
    return result
"""
特別パッチが必要となる日が存在しないかチェックして結果をresultに入れる関数
"""
def SPC_PATCH_CHK(sheet_name):
	result=[]
	ws = wb[sheet_name]
	max_R=ws.max_row
	max_C=ws.max_column
	cnt = 0
	print(sheet_name+' の処理を開始')
	for i in range(4,max_C):
		chk_cell = i -1
		for j in range(6,max_R):
			if ws.cell(row=j, column=chk_cell).value - ws.cell(row=j, column=i).value >= 2:
				result.append(
				str(format(ws.cell(row=1,column=i).value,'%Y/%m/%d %H:%M:%S')) +'\t'
				+ ws.cell(row=j,column=1).value +'\t'
				+ prfc[str(ws.cell(row=j, column=1).value)] +'\t'
				+ str(ws.cell(row=j, column=chk_cell).value - ws.cell(row=j, column=i).value)
				+ '\n'
				)
				cnt += 1
	print(sheet_name +' の処理を完了')
	if cnt == 0:
		print(sheet_name+' には特別パッチが必要なレコードはありませんでした。\n')
	else:
		print(sheet_name+' の特別パッチ対象となるレコードは'+ str(cnt) +'件でした。\n')
	return result

"""
都道府県コードを辞書に登録
"""
prfc={"北海道":"01","青森":"02","岩手":"03","宮城":"04","秋田":"05","山形":"06","福島":"07","茨城":"08","栃木":"09","群馬":"10","埼玉":"11",
	"千葉":"12","東京":"13","神奈川":"14","新潟":"15","富山":"16","石川":"17","福井":"18","山梨":"19","長野":"20","岐阜":"21","静岡":"22",
	"愛知":"23","三重":"24","滋賀":"25","京都":"26","大阪":"27","兵庫":"28","奈良":"29","和歌山":"30","鳥取":"31","島根":"32","岡山":"33",
	"広島":"34","山口":"35","徳島":"36","香川":"37","愛媛":"38","高知":"39","福岡":"40","佐賀":"41","長崎":"42","熊本":"43","大分":"44",
	"宮崎":"45","鹿児島":"46","沖縄":"47"}


"""★★★★★★★★★処理ここから★★★★★★★★★★"""
"""------------------------------------------------------
作業ディレクトリの確認:今はデスクトップ直下の「python￥wrk」で固定にしてる。
------------------------------------------------------"""
desktop_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Desktop"
kakunin_dir = desktop_path+'\python\wrk'
if os.path.exists(kakunin_dir) == False:
    print("作業ディレクトリ "+kakunin_dir+" が存在しません。\n処理を終了します。")
    sys.exit(1)
os.chdir(kakunin_dir)
wrk_dir = os.getcwd()

"""------------------------------------------------------
SQLの材料フィルを見つけて開く
openpexlではxlsxファイルじゃないと扱えないので拡張子は固定
------------------------------------------------------"""
path = pathlib.Path(wrk_dir)
for pass_obj in path.iterdir():
    if pass_obj.match("*D+N日表記*.xlsx"):
            wb = openpyxl.load_workbook(pass_obj,data_only=True) #data~で関数が入ったセルも値だけ持ってくるようにする
print(str(pass_obj) + " に対して処理を行います。\n")

"""------------------------------------------------------
休配日パッチ材料ファイル作成
------------------------------------------------------"""
"SEJ店舗のチェック"
print("休配日パッチ材料ファイルの作成を開始します。")
with open('SEJ_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as sej:
    writer = csv.writer(sej)
    for initem in READ_LDTM('ＳＥＪ店舗 (差分)'):
        sej.write(initem)
    sej.close()
"その他のチェック"
with open('OTHER_INSERT.csv', 'w', encoding = 'utf_8', newline = '\n') as oth:
    writer = csv.writer(oth)
    for initem in READ_LDTM('その他 (差分)'):
        oth.write(initem)
    oth.close()
print("「休配日パッチ材料ファイル」の作成が完了しました。\n\n")


"""------------------------------------------------------
特別パッチの要否確認
------------------------------------------------------"""
"SEJ店舗のチェック"
print("特別パッチの要否チェックを開始します。")
with open('SPCL_PATCH_CHK_SEJ.csv', 'w', encoding = 'utf_8', newline = '\n') as sej_sp:
	writer = csv.writer(sej_sp)
	for initem in SPC_PATCH_CHK('ＳＥＪ店舗 (差分)'):
		sej_sp.write(initem)
	sej_sp.close()
"その他のチェック"
with open('SPCL_PATCH_CHK_OTEHR.csv', 'w', encoding = 'utf_8', newline = '\n') as oth_sp:
	writer = csv.writer(oth_sp)
	for initem in SPC_PATCH_CHK('その他 (差分)'):
		oth_sp.write(initem)
	oth_sp.close()

print("特別パッチの要否チェックは全て完了しました\n\n。")
print("全ての処理が完了しました。")
